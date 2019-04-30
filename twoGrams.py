from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


def sorVec(phrase):
    alp = 'abcdefghijklmnopqrstuvwxyz'
    table = {}
    for i in range(26):
        for j in range(26):
            table[ alp[ i ] + alp[ j ] ] = 0

    list = [ ]
    for i in range(len(phrase) - 1):
        list.append(phrase[ i ] + phrase[ i + 1 ])

    for i in range(len(list)):
        if list[ i ] in table:
            table[ list[ i ] ] += 1

    return table.values(), table.keys()


def creatXL():
    init = "init"
    _, keys = sorVec(init)
    keys = list(keys)
    book = Workbook()
    ws = book.active
    sheet1 = book.create_sheet("2Grams")
    sheet1.cell(1, 1, "Phrase/Dic")

    for i in range(26 * 26):
        sheet1.cell(row=1, column=i + 2, value=keys[ i ])

    book.save("2Grams.xlsx")


def insertData(pharse, xls):
    wb = load_workbook(xls)
    ws = wb[ "2Grams" ]
    ws.calculate_dimension()
    row = ws.max_row
    ws.cell(row + 1, 1, pharse)
    values, _ = sorVec(pharse)
    values = list(values)
    for i in range(26 * 26):
        ws.cell(row + 1, i + 2, values[ i ])

    wb.save(xls)


def twoGrams(listPhrase):
    for i in range(len(listPhrase)):
        insertData(listPhrase[ i ], "2Grams.xlsx")


if __name__ == "__main__":
    wbData = load_workbook("CleanPublication.xlsx")
    wsData = wbData[ "Feuil1" ]
    wsData.calculate_dimension()
    titles = [ ]
    for row in wsData.iter_rows('D{}:D{}'.format(wsData.min_row, wsData.max_row)):
        for cell in row:
            titles.append(cell.value)

    twoGrams(titles)
